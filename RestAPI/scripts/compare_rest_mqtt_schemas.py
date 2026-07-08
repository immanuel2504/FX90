#!/usr/bin/env python3
"""Compare REST (openAPISpec2.yaml) field schemas against MQTT expanded schemas."""

from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
REST_SPEC = ROOT / "openAPISpec2.yaml"
MAP_JSON = ROOT / "RestAPI" / "mappings" / "rest_mqtt_map.json"
CMD_EXPANDED = ROOT / "schemas" / "commands_expanded"
RESP_EXPANDED = ROOT / "schemas" / "response_expanded"
CMD_SPLIT = ROOT / "schemas" / "commands"
RESP_SPLIT = ROOT / "schemas" / "response"
OUTPUT = ROOT / "RestAPI" / "rest_vs_mqtt_field_report.xlsx"

# When True, compare against the split + references source of truth (what the
# doc build uses) instead of the stale *_expanded copies.
USE_SPLIT_SOURCE = os.environ.get("MQTT_SOURCE", "split").lower() == "split"

# REST path parameter names mapped to MQTT payload field names. MQTT has no path
# parameters, so these fields appear MQTT-only but are required for parity.
PATH_PARAM_MQTT_ALIASES: dict[str, str] = {
    "certname": "name",
}

# Per-endpoint overrides when the MQTT payload uses a different field name than REST.
PATH_PARAM_MQTT_ALIASES_BY_ENDPOINT: dict[tuple[str, str], dict[str, str]] = {
    ("PUT", "/cloud/apps/{appname}/pass-through"): {"appname": "userapp"},
}


def mqtt_fields_from_path_params(
    path_parameters: list[dict[str, Any]],
    *,
    method: str = "",
    path: str = "",
) -> set[str]:
    """MQTT payload fields that correspond to REST path parameters."""
    endpoint_aliases = PATH_PARAM_MQTT_ALIASES_BY_ENDPOINT.get((method.upper(), path), {})
    out: set[str] = set()
    for param in path_parameters:
        pname = param.get("name")
        if not pname:
            continue
        out.add(endpoint_aliases.get(pname, PATH_PARAM_MQTT_ALIASES.get(pname, pname)))
    return out

MQTT_RE = re.compile(r"MQTT API\s*:-\s*(\S+)", re.IGNORECASE)
HTTP_METHODS = {"get", "put", "post", "delete", "patch"}


def load_yaml(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def normalize_desc(text: str | None) -> str:
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text.strip().lower())
    return text


def normalize_type(t: str | list[str] | None) -> str:
    if t is None:
        return ""
    if isinstance(t, list):
        return "|".join(sorted(normalize_type(x) for x in t if x))
    mapping = {"integer": "number", "int": "number", "float": "number", "double": "number"}
    return mapping.get(t.lower(), t.lower())


def normalize_enum(values: list[Any] | None) -> list[str]:
    if not values:
        return []
    out: list[str] = []
    for v in values:
        if isinstance(v, bool):
            out.append(str(v).lower())
        else:
            out.append(str(v).strip().lower())
    return sorted(set(out))


def resolve_ref(ref: str, root: dict[str, Any]) -> dict[str, Any]:
    if not ref.startswith("#/"):
        return {}
    parts = ref.lstrip("#/").split("/")
    node: Any = root
    for part in parts:
        if not isinstance(node, dict) or part not in node:
            return {}
        node = node[part]
    return node if isinstance(node, dict) else {}


def resolve_schema(schema: Any, root: dict[str, Any], seen: set[str] | None = None) -> dict[str, Any]:
    if not isinstance(schema, dict):
        return {}
    if seen is None:
        seen = set()
    if "$ref" in schema:
        ref = schema["$ref"]
        if ref in seen:
            return {}
        seen = set(seen)
        seen.add(ref)
        resolved = resolve_schema(resolve_ref(ref, root), root, seen)
        merged = {k: v for k, v in schema.items() if k != "$ref"}
        if merged:
            return deep_merge(resolved, merged)
        return resolved
    out = dict(schema)
    for key in ("items", "additionalProperties"):
        if key in out and isinstance(out[key], dict):
            out[key] = resolve_schema(out[key], root, seen)
    if "properties" in out and isinstance(out["properties"], dict):
        out["properties"] = {
            k: resolve_schema(v, root, seen) for k, v in out["properties"].items()
        }
    for combiner in ("allOf", "anyOf", "oneOf"):
        if combiner in out and isinstance(out[combiner], list):
            out[combiner] = [resolve_schema(s, root, seen) for s in out[combiner]]
    return out


def deep_merge(a: dict[str, Any], b: dict[str, Any]) -> dict[str, Any]:
    result = dict(a)
    for k, v in b.items():
        if k in result and isinstance(result[k], dict) and isinstance(v, dict):
            result[k] = deep_merge(result[k], v)
        else:
            result[k] = v
    return result


def merge_all_of(schemas: list[dict[str, Any]]) -> dict[str, Any]:
    merged: dict[str, Any] = {"type": "object", "properties": {}, "required": []}
    req: set[str] = set()
    for schema in schemas:
        if not schema:
            continue
        if schema.get("type") == "object" or "properties" in schema:
            merged["properties"].update(schema.get("properties") or {})
            req.update(schema.get("required") or [])
        for k, v in schema.items():
            if k not in ("properties", "required", "allOf", "oneOf", "anyOf"):
                merged[k] = v
    if req:
        merged["required"] = sorted(req)
    return merged


def flatten_schema_variants(schema: dict[str, Any]) -> list[dict[str, Any]]:
    if not schema:
        return [{}]
    if "allOf" in schema:
        return [merge_all_of(schema["allOf"])]
    if "oneOf" in schema:
        return [s for s in schema["oneOf"] if isinstance(s, dict)]
    if "anyOf" in schema:
        return [s for s in schema["anyOf"] if isinstance(s, dict)]
    return [schema]


def collect_fields(
    schema: dict[str, Any],
    prefix: str = "",
    required: set[str] | None = None,
    fields: dict[str, dict[str, Any]] | None = None,
) -> dict[str, dict[str, Any]]:
    if fields is None:
        fields = {}
    if required is None:
        required = set()

    schema_type = schema.get("type")
    props = schema.get("properties") or {}
    local_required = set(schema.get("required") or []) | required

    if schema_type == "array" or "items" in schema:
        items = schema.get("items") or {}
        if isinstance(items, dict):
            collect_fields(items, f"{prefix}[]" if prefix else "[]", set(), fields)
        return fields

    if schema.get("patternProperties"):
        for pat, sub in schema["patternProperties"].items():
            if isinstance(sub, dict):
                collect_fields(sub, f"{prefix}[{pat}]" if prefix else f"[{pat}]", set(), fields)
        return fields

    if schema_type == "object" or props or local_required:
        for name, sub in props.items():
            path = f"{prefix}.{name}" if prefix else name
            entry = fields.setdefault(
                path,
                {
                    "type": "",
                    "enum": [],
                    "description": "",
                    "required_rest": False,
                    "required_mqtt": False,
                },
            )
            if isinstance(sub, dict):
                if sub.get("type"):
                    entry["type"] = normalize_type(sub.get("type"))
                if sub.get("enum"):
                    entry["enum"] = normalize_enum(sub.get("enum"))
                if sub.get("description"):
                    entry["description"] = sub.get("description", "")
                if name in local_required:
                    pass  # caller sets required_* flags
                nested_type = sub.get("type")
                nested_props = sub.get("properties")
                if nested_type == "object" or nested_props or sub.get("oneOf") or sub.get("allOf"):
                    for variant in flatten_schema_variants(sub):
                        child_required = set(variant.get("required") or [])
                        if name in local_required:
                            child_required = child_required  # noqa: B018
                        collect_fields(
                            variant,
                            path,
                            child_required if name in local_required else set(),
                            fields,
                        )
        return fields

    if prefix:
        entry = fields.setdefault(
            prefix,
            {"type": "", "enum": [], "description": "", "required_rest": False, "required_mqtt": False},
        )
        if schema.get("type"):
            entry["type"] = normalize_type(schema.get("type"))
        if schema.get("enum"):
            entry["enum"] = normalize_enum(schema.get("enum"))
        if schema.get("description"):
            entry["description"] = schema.get("description", "")
    return fields


def collect_fields_from_variants(schema: dict[str, Any]) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    for variant in flatten_schema_variants(schema):
        partial = collect_fields(variant)
        req = set(variant.get("required") or [])
        for path, meta in partial.items():
            top = path.split(".")[0].split("[")[0]
            target = fields.setdefault(path, {"type": "", "enum": [], "description": ""})
            if meta.get("type"):
                target["type"] = meta["type"]
            if meta.get("enum"):
                target["enum"] = sorted(set(target.get("enum", [])) | set(meta["enum"]))
            if meta.get("description") and not target.get("description"):
                target["description"] = meta["description"]
            if top in req:
                target["required"] = True
    return fields


def mark_required(fields: dict[str, dict[str, Any]], schema: dict[str, Any], flag: str) -> None:
    for variant in flatten_schema_variants(schema):
        req = set(variant.get("required") or [])
        for path in fields:
            top = path.split(".")[0].split("[")[0]
            if top in req:
                fields[path][flag] = True


def extract_rest_request_schema(operation: dict[str, Any], root: dict[str, Any]) -> dict[str, Any]:
    body = operation.get("requestBody") or {}
    content = body.get("content") or {}
    json_content = content.get("application/json") or {}
    schema = json_content.get("schema") or {}
    resolved = resolve_schema(schema, root)

    params_schema: dict[str, Any] = {"type": "object", "properties": {}, "required": []}
    for param in operation.get("parameters") or []:
        if param.get("in") in ("query", "path"):
            name = param.get("name")
            if not name:
                continue
            params_schema["properties"][name] = {
                "type": param.get("schema", {}).get("type") or param.get("type", "string"),
                "description": param.get("description", ""),
                "enum": param.get("schema", {}).get("enum") or param.get("enum"),
            }
            if param.get("required"):
                params_schema["required"].append(name)

    if params_schema["properties"] and resolved:
        merged = {
            "type": "object",
            "properties": {**params_schema["properties"], **(resolved.get("properties") or {})},
            "required": sorted(set(params_schema.get("required") or []) | set(resolved.get("required") or [])),
        }
        for combiner in ("oneOf", "allOf", "anyOf"):
            if combiner in resolved:
                merged[combiner] = resolved[combiner]
        return merged
    if params_schema["properties"]:
        return params_schema
    return resolved


def extract_rest_response_schema(operation: dict[str, Any], root: dict[str, Any]) -> dict[str, Any]:
    responses = operation.get("responses") or {}
    for code in ("200", "201", "204", "default"):
        if code not in responses:
            continue
        content = (responses[code].get("content") or {}).get("application/json") or {}
        schema = content.get("schema")
        if schema:
            return resolve_schema(schema, root)
    return {}


def mqtt_success_payload(response_schema: dict[str, Any]) -> dict[str, Any]:
    payload = (response_schema.get("properties") or {}).get("payload") or {}
    if "oneOf" in payload:
        for variant in payload["oneOf"]:
            title = (variant.get("title") or "").lower()
            if title in ("success", "") and "code" not in (variant.get("properties") or {}):
                return variant
        for variant in payload["oneOf"]:
            if "code" not in (variant.get("properties") or {}):
                return variant
        return payload["oneOf"][0]
    return payload


def find_mqtt_file(base_dir: Path, mqtt_command: str, suffix: str = "") -> Path | None:
    stem = f"{mqtt_command}{suffix}"
    matches = list(base_dir.rglob(f"{stem}.json"))
    if matches:
        return matches[0]
    alt = mqtt_command.replace("-", "_")
    matches = list(base_dir.rglob(f"{alt}{suffix}.json"))
    return matches[0] if matches else None


def _load_structured_file(path: Path) -> Any:
    with path.open(encoding="utf-8-sig") as f:
        if path.suffix.lower() in (".yaml", ".yml"):
            return yaml.safe_load(f)
        return json.load(f)


def resolve_file_refs(node: Any, base_file: Path, seen: set[Path] | None = None) -> Any:
    """Recursively resolve relative file $refs used by the split MQTT schemas."""
    if seen is None:
        seen = set()
    if isinstance(node, list):
        return [resolve_file_refs(n, base_file, seen) for n in node]
    if not isinstance(node, dict):
        return node
    if "$ref" in node and isinstance(node["$ref"], str) and not node["$ref"].startswith("#/"):
        target = (base_file.parent / node["$ref"]).resolve()
        extras = {k: v for k, v in node.items() if k != "$ref"}
        if not target.is_file() or target in seen:
            return resolve_file_refs(extras, base_file, seen) if extras else {}
        sub = _load_structured_file(target)
        resolved = resolve_file_refs(sub, target, seen | {target})
        if isinstance(resolved, dict) and extras:
            merged = dict(resolved)
            merged.update(resolve_file_refs(extras, base_file, seen))
            return merged
        return resolved
    return {k: resolve_file_refs(v, base_file, seen) for k, v in node.items()}


def load_mqtt_command_payload(mqtt_command: str) -> dict[str, Any] | None:
    if USE_SPLIT_SOURCE:
        path = find_mqtt_file(CMD_SPLIT, mqtt_command)
        if not path:
            return None
        data = resolve_file_refs(_load_structured_file(path), path)
        return (data.get("properties") or {}).get("payload") or {}
    path = find_mqtt_file(CMD_EXPANDED, mqtt_command)
    if not path:
        return None
    data = json.loads(path.read_text(encoding="utf-8"))
    return (data.get("properties") or {}).get("payload") or {}


def load_mqtt_response_payload(mqtt_command: str) -> dict[str, Any] | None:
    if USE_SPLIT_SOURCE:
        path = find_mqtt_file(RESP_SPLIT, mqtt_command, "_response")
        if not path:
            return None
        data = resolve_file_refs(_load_structured_file(path), path)
        return mqtt_success_payload(data)
    path = find_mqtt_file(RESP_EXPANDED, mqtt_command, "_response")
    if not path:
        return None
    data = json.loads(path.read_text(encoding="utf-8"))
    return mqtt_success_payload(data)


def compare_field_sets(
    rest_fields: dict[str, dict[str, Any]],
    mqtt_fields: dict[str, dict[str, Any]],
    *,
    rest_required_flag: str,
    mqtt_required_flag: str,
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    all_paths = sorted(set(rest_fields) | set(mqtt_fields))

    for path in all_paths:
        r = rest_fields.get(path, {})
        m = mqtt_fields.get(path, {})

        if path not in rest_fields:
            issues.append(
                {
                    "field": path,
                    "issue": "MQTT only field",
                    "rest": "",
                    "mqtt": summarize_field(m),
                    "severity": "high",
                }
            )
            continue
        if path not in mqtt_fields:
            issues.append(
                {
                    "field": path,
                    "issue": "REST only field",
                    "rest": summarize_field(r),
                    "mqtt": "",
                    "severity": "high",
                }
            )
            continue

        r_type = r.get("type", "")
        m_type = m.get("type", "")
        if r_type and m_type and r_type != m_type:
            issues.append(
                {
                    "field": path,
                    "issue": "Type mismatch",
                    "rest": r_type,
                    "mqtt": m_type,
                    "severity": "high",
                }
            )

        r_enum = normalize_enum(r.get("enum"))
        m_enum = normalize_enum(m.get("enum"))
        if r_enum or m_enum:
            if r_enum != m_enum:
                missing_rest = sorted(set(m_enum) - set(r_enum))
                missing_mqtt = sorted(set(r_enum) - set(m_enum))
                issues.append(
                    {
                        "field": path,
                        "issue": "Enum mismatch",
                        "rest": ", ".join(r_enum) if r_enum else "(none)",
                        "mqtt": ", ".join(m_enum) if m_enum else "(none)",
                        "severity": "critical",
                        "detail": f"REST missing: {missing_mqtt or '-'}; MQTT missing: {missing_rest or '-'}",
                    }
                )

        r_desc = normalize_desc(r.get("description"))
        m_desc = normalize_desc(m.get("description"))
        if r_desc and m_desc and r_desc != m_desc:
            issues.append(
                {
                    "field": path,
                    "issue": "Description mismatch",
                    "rest": r.get("description", ""),
                    "mqtt": m.get("description", ""),
                    "severity": "medium",
                }
            )
        elif r_desc and not m_desc:
            issues.append(
                {
                    "field": path,
                    "issue": "Description missing in MQTT",
                    "rest": r.get("description", ""),
                    "mqtt": "",
                    "severity": "low",
                }
            )
        elif m_desc and not r_desc:
            issues.append(
                {
                    "field": path,
                    "issue": "Description missing in REST",
                    "rest": "",
                    "mqtt": m.get("description", ""),
                    "severity": "low",
                }
            )

        r_req = r.get(rest_required_flag, False)
        m_req = m.get(mqtt_required_flag, False)
        if r_req != m_req:
            issues.append(
                {
                    "field": path,
                    "issue": "Required mismatch",
                    "rest": "required" if r_req else "optional",
                    "mqtt": "required" if m_req else "optional",
                    "severity": "high",
                }
            )

    return issues


def summarize_field(meta: dict[str, Any]) -> str:
    parts = []
    if meta.get("type"):
        parts.append(meta["type"])
    if meta.get("enum"):
        parts.append("enum=" + ",".join(normalize_enum(meta["enum"])))
    return "; ".join(parts) if parts else "(no type)"


def build_rest_operations(spec: dict[str, Any], mapping: dict[str, Any]) -> list[dict[str, Any]]:
    map_by_path: dict[tuple[str, str], dict[str, Any]] = {}
    for item in mapping.get("restToMqtt", []):
        map_by_path[(item["httpMethod"].upper(), item["restPath"])] = item

    ops: list[dict[str, Any]] = []
    for path, path_item in (spec.get("paths") or {}).items():
        for method, operation in path_item.items():
            if method.lower() not in HTTP_METHODS:
                continue
            desc = operation.get("description") or ""
            mqtt_match = MQTT_RE.search(desc)
            mqtt_from_desc = mqtt_match.group(1) if mqtt_match else ""
            mapped = map_by_path.get((method.upper(), path), {})
            mqtt_command = mapped.get("mqttCommand") or mqtt_from_desc
            path_params = [
                p for p in (path_item.get("parameters") or []) if p.get("in") == "path"
            ]
            ops.append(
                {
                    "path": path,
                    "method": method.upper(),
                    "operationId": operation.get("operationId", ""),
                    "mqtt_command": mqtt_command,
                    "operation": operation,
                    "path_parameters": path_params,
                    "mapped": bool(mapped),
                }
            )
    return ops


def format_enum(values: list[Any] | None) -> str:
    normalized = normalize_enum(values)
    return ", ".join(normalized) if normalized else ""


def build_field_alignment_rows(
    *,
    endpoint: str,
    operation_id: str,
    mqtt_command: str,
    side: str,
    rest_fields: dict[str, dict[str, Any]],
    mqtt_fields: dict[str, dict[str, Any]],
    issues: list[dict[str, Any]],
    path_mqtt_fields: set[str] | None = None,
) -> list[dict[str, Any]]:
    """Build per-field alignment rows for the full comparison report."""
    path_mqtt_fields = path_mqtt_fields or set()
    issues_by_field: dict[str, list[str]] = defaultdict(list)
    for issue in issues:
        issues_by_field[issue["field"]].append(issue["issue"])

    rows: list[dict[str, Any]] = []
    for field in sorted(set(rest_fields) | set(mqtt_fields)):
        r = rest_fields.get(field, {})
        m = mqtt_fields.get(field, {})
        in_rest = field in rest_fields
        in_mqtt = field in mqtt_fields

        r_type = normalize_type(r.get("type"))
        m_type = normalize_type(m.get("type"))
        r_enum = normalize_enum(r.get("enum"))
        m_enum = normalize_enum(m.get("enum"))
        r_desc = normalize_desc(r.get("description"))
        m_desc = normalize_desc(m.get("description"))
        r_req = bool(r.get("required_rest", False))
        m_req = bool(m.get("required_mqtt", False))

        type_match = (not r_type and not m_type) or (r_type == m_type)
        enum_match = r_enum == m_enum
        desc_match = (not r_desc and not m_desc) or (r_desc == m_desc)
        required_match = r_req == m_req

        field_issues = list(issues_by_field.get(field, []))
        if not in_rest and in_mqtt and field in path_mqtt_fields:
            status = "OK (MQTT path-param field)"
        elif not field_issues:
            status = "OK"
        else:
            status = "; ".join(sorted(set(field_issues)))

        rows.append(
            {
                "endpoint": endpoint,
                "operationId": operation_id,
                "mqtt_command": mqtt_command,
                "side": side,
                "field": field,
                "in_rest": "yes" if in_rest else "no",
                "in_mqtt": "yes" if in_mqtt else "no",
                "rest_type": r_type,
                "mqtt_type": m_type,
                "type_match": "yes" if type_match else "no",
                "rest_enum": format_enum(r.get("enum")),
                "mqtt_enum": format_enum(m.get("enum")),
                "enum_match": "yes" if enum_match else "no",
                "rest_description": r.get("description", "") or "",
                "mqtt_description": m.get("description", "") or "",
                "description_match": "yes" if desc_match else "no",
                "rest_required": "required" if r_req else "optional",
                "mqtt_required": "required" if m_req else "optional",
                "required_match": "yes" if required_match else "no",
                "status": status,
            }
        )
    return rows


def compare_operation(
    op: dict[str, Any], root: dict[str, Any]
) -> tuple[list[dict[str, Any]], str | None, list[dict[str, Any]]]:
    mqtt_command = op["mqtt_command"]
    if not mqtt_command:
        return [], "No MQTT command identified", []

    method = op["method"]
    operation = op["operation"]
    rows: list[dict[str, Any]] = []
    note_parts: list[str] = []
    path_mqtt_fields: set[str] = set()

    if method in ("PUT", "POST", "PATCH", "DELETE"):
        rest_schema = extract_rest_request_schema(operation, root)
        mqtt_schema = load_mqtt_command_payload(mqtt_command)
        if mqtt_schema is None:
            return [], f"MQTT command schema not found for {mqtt_command}", []
        if not rest_schema and not mqtt_schema.get("properties"):
            return [], "Both sides empty request payload", []
        rest_fields = collect_fields_from_variants(rest_schema)
        mark_required(rest_fields, rest_schema, "required_rest")
        mqtt_fields = collect_fields_from_variants(mqtt_schema)
        mark_required(mqtt_fields, mqtt_schema, "required_mqtt")
        side = "request"
    else:
        rest_schema = extract_rest_response_schema(operation, root)
        mqtt_schema = load_mqtt_response_payload(mqtt_command)
        if mqtt_schema is None:
            return [], f"MQTT response schema not found for {mqtt_command}", []
        rest_fields = collect_fields_from_variants(rest_schema)
        mark_required(rest_fields, rest_schema, "required_rest")
        mqtt_fields = collect_fields_from_variants(mqtt_schema)
        mark_required(mqtt_fields, mqtt_schema, "required_mqtt")
        side = "response"

    issues = compare_field_sets(
        rest_fields,
        mqtt_fields,
        rest_required_flag="required_rest",
        mqtt_required_flag="required_mqtt",
    )
    if method in ("PUT", "POST", "PATCH", "DELETE"):
        path_mqtt_fields = mqtt_fields_from_path_params(
            op.get("path_parameters") or [],
            method=method,
            path=op["path"],
        )
        mqtt_only_paths = set(mqtt_fields) - set(rest_fields)
        expected_path_fields = sorted(mqtt_only_paths & path_mqtt_fields)
        if expected_path_fields:
            note_parts.append(
                "MQTT-only path-param payload fields (expected): "
                + ", ".join(expected_path_fields)
            )
        issues = [
            i
            for i in issues
            if not (
                i["issue"] in ("MQTT only field", "Required mismatch")
                and i["field"] in path_mqtt_fields
            )
        ]

    endpoint = f"{method} {op['path']}"
    alignment_rows = build_field_alignment_rows(
        endpoint=endpoint,
        operation_id=op["operationId"],
        mqtt_command=mqtt_command,
        side=side,
        rest_fields=rest_fields,
        mqtt_fields=mqtt_fields,
        issues=issues,
        path_mqtt_fields=path_mqtt_fields,
    )
    for issue in issues:
        rows.append(
            {
                "endpoint": endpoint,
                "operationId": op["operationId"],
                "mqtt_command": mqtt_command,
                "side": side,
                **issue,
            }
        )
    if not rows and not rest_fields and not mqtt_fields:
        note_parts.append("No comparable fields extracted")
    note = "; ".join(note_parts) if note_parts else None
    return rows, note, alignment_rows


def autosize_columns(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 80)
        ws.column_dimensions[letter].width = width


def write_report(
    all_rows: list[dict[str, Any]],
    ops: list[dict[str, Any]],
    notes: list[dict[str, Any]],
    alignment_rows: list[dict[str, Any]],
) -> None:
    wb = openpyxl.Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    headers = ["Metric", "Count"]
    ws_sum.append(headers)
    for cell in ws_sum[1]:
        cell.font = Font(bold=True)

    severity_counts = defaultdict(int)
    issue_counts = defaultdict(int)
    endpoint_issues = defaultdict(int)
    for row in all_rows:
        severity_counts[row["severity"]] += 1
        issue_counts[row["issue"]] += 1
        endpoint_issues[row["endpoint"]] += 1

    matched = sum(1 for op in ops if op["mqtt_command"])
    no_mqtt = [op for op in ops if op["mqtt_command"] and any(n.get("endpoint") == f"{op['method']} {op['path']}" and "not found" in (n.get("note") or "") for n in notes)]

    aligned_ok = sum(1 for r in alignment_rows if r["status"] == "OK")
    aligned_path = sum(1 for r in alignment_rows if r["status"] == "OK (MQTT path-param field)")
    aligned_issues = sum(1 for r in alignment_rows if r["status"] not in ("OK", "OK (MQTT path-param field)"))
    enum_compared = sum(1 for r in alignment_rows if r["rest_enum"] or r["mqtt_enum"])
    enum_match = sum(1 for r in alignment_rows if (r["rest_enum"] or r["mqtt_enum"]) and r["enum_match"] == "yes")
    desc_compared = sum(1 for r in alignment_rows if r["rest_description"] or r["mqtt_description"])
    desc_match = sum(1 for r in alignment_rows if (r["rest_description"] or r["mqtt_description"]) and r["description_match"] == "yes")

    summary_rows = [
        ("REST operations in openAPISpec2", len(ops)),
        ("Operations with MQTT command label", matched),
        ("Operations with field issues", len(endpoint_issues)),
        ("Operations missing MQTT schema", len(no_mqtt)),
        ("Total field issues", len(all_rows)),
        ("Fields compared (REST vs MQTT)", len(alignment_rows)),
        ("Fields fully aligned (OK)", aligned_ok),
        ("Fields aligned as MQTT path-param", aligned_path),
        ("Fields with mismatches", aligned_issues),
        ("Enum fields compared", enum_compared),
        ("Enum fields matching", enum_match),
        ("Description fields compared", desc_compared),
        ("Description fields matching", desc_match),
        ("Critical (enum)", severity_counts["critical"]),
        ("High (field/type/required)", severity_counts["high"]),
        ("Medium (description)", severity_counts["medium"]),
        ("Low (missing description)", severity_counts["low"]),
    ]
    for metric, count in summary_rows:
        ws_sum.append([metric, count])

    ws_sum.append([])
    ws_sum.append(["Issue type", "Count"])
    ws_sum[ws_sum.max_row][0].font = Font(bold=True)
    ws_sum[ws_sum.max_row][1].font = Font(bold=True)
    for issue, count in sorted(issue_counts.items(), key=lambda x: (-x[1], x[0])):
        ws_sum.append([issue, count])

    # All issues
    ws = wb.create_sheet("All Issues")
    issue_headers = [
        "Severity",
        "Issue",
        "Endpoint",
        "operationId",
        "MQTT command",
        "Side",
        "Field",
        "REST value",
        "MQTT value",
        "Detail",
    ]
    ws.append(issue_headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    severity_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    for row in sorted(all_rows, key=lambda r: (severity_order.get(r["severity"], 9), r["endpoint"], r["field"])):
        ws.append(
            [
                row["severity"],
                row["issue"],
                row["endpoint"],
                row["operationId"],
                row["mqtt_command"],
                row["side"],
                row["field"],
                row.get("rest", ""),
                row.get("mqtt", ""),
                row.get("detail", ""),
            ]
        )
    autosize_columns(ws)

    # Endpoint summary (all compared endpoints)
    ws_ep_sum = wb.create_sheet("Endpoint Summary")
    ep_headers = [
        "Endpoint",
        "MQTT command",
        "Side",
        "Fields compared",
        "Fields OK",
        "Path-param fields",
        "Field mismatches",
        "Enum mismatches",
        "Description mismatches",
        "Type mismatches",
        "Required mismatches",
        "Overall status",
    ]
    ws_ep_sum.append(ep_headers)
    for cell in ws_ep_sum[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    by_ep_align: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in alignment_rows:
        key = (row["endpoint"], row["mqtt_command"], row["side"])
        by_ep_align[key].append(row)

    for (endpoint, mqtt_command, side), rows in sorted(by_ep_align.items()):
        ok = sum(1 for r in rows if r["status"] == "OK")
        path_ok = sum(1 for r in rows if r["status"] == "OK (MQTT path-param field)")
        mismatches = [r for r in rows if r["status"] not in ("OK", "OK (MQTT path-param field)")]
        enum_mm = sum(1 for r in rows if r["enum_match"] == "no" and (r["rest_enum"] or r["mqtt_enum"]))
        desc_mm = sum(1 for r in rows if r["description_match"] == "no" and (r["rest_description"] or r["mqtt_description"]))
        type_mm = sum(1 for r in rows if r["type_match"] == "no" and (r["rest_type"] or r["mqtt_type"]))
        req_mm = sum(1 for r in rows if r["required_match"] == "no")
        overall = "ALIGNED" if not mismatches else "MISMATCH"
        ws_ep_sum.append(
            [
                endpoint,
                mqtt_command,
                side,
                len(rows),
                ok,
                path_ok,
                len(mismatches),
                enum_mm,
                desc_mm,
                type_mm,
                req_mm,
                overall,
            ]
        )
    autosize_columns(ws_ep_sum)

    # Field-by-field alignment
    ws_align = wb.create_sheet("Field Alignment")
    align_headers = [
        "Endpoint",
        "MQTT command",
        "Side",
        "Field",
        "In REST",
        "In MQTT",
        "REST type",
        "MQTT type",
        "Type match",
        "REST enum",
        "MQTT enum",
        "Enum match",
        "REST description",
        "MQTT description",
        "Description match",
        "REST required",
        "MQTT required",
        "Required match",
        "Status",
    ]
    ws_align.append(align_headers)
    for cell in ws_align[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    path_fill = PatternFill("solid", fgColor="FFEB9C")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")

    for row in sorted(alignment_rows, key=lambda r: (r["endpoint"], r["field"])):
        ws_align.append(
            [
                row["endpoint"],
                row["mqtt_command"],
                row["side"],
                row["field"],
                row["in_rest"],
                row["in_mqtt"],
                row["rest_type"],
                row["mqtt_type"],
                row["type_match"],
                row["rest_enum"],
                row["mqtt_enum"],
                row["enum_match"],
                row["rest_description"],
                row["mqtt_description"],
                row["description_match"],
                row["rest_required"],
                row["mqtt_required"],
                row["required_match"],
                row["status"],
            ]
        )
        status_cell = ws_align.cell(row=ws_align.max_row, column=len(align_headers))
        if row["status"] == "OK":
            status_cell.fill = ok_fill
        elif row["status"] == "OK (MQTT path-param field)":
            status_cell.fill = path_fill
        else:
            status_cell.fill = bad_fill
    autosize_columns(ws_align)

    # By endpoint pivot-like sheet
    ws_ep = wb.create_sheet("By Endpoint")
    ws_ep.append(["Endpoint", "MQTT command", "Issues", "Critical", "High", "Medium", "Low", "Top issues"])
    for cell in ws_ep[1]:
        cell.font = Font(bold=True)

    by_ep: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in all_rows:
        by_ep[row["endpoint"]].append(row)

    for endpoint in sorted(by_ep):
        rows = by_ep[endpoint]
        counts = defaultdict(int)
        for r in rows:
            counts[r["severity"]] += 1
        top = ", ".join(sorted({r["issue"] for r in rows}))
        ws_ep.append(
            [
                endpoint,
                rows[0]["mqtt_command"],
                len(rows),
                counts["critical"],
                counts["high"],
                counts["medium"],
                counts["low"],
                top,
            ]
        )
    autosize_columns(ws_ep)

    # Enum mismatches only
    ws_enum = wb.create_sheet("Enum Mismatches")
    ws_enum.append(issue_headers)
    for cell in ws_enum[1]:
        cell.font = Font(bold=True)
    for row in all_rows:
        if row["issue"] == "Enum mismatch":
            ws_enum.append(
                [
                    row["severity"],
                    row["issue"],
                    row["endpoint"],
                    row["operationId"],
                    row["mqtt_command"],
                    row["side"],
                    row["field"],
                    row.get("rest", ""),
                    row.get("mqtt", ""),
                    row.get("detail", ""),
                ]
            )
    autosize_columns(ws_enum)

    # Description mismatches only
    ws_desc = wb.create_sheet("Description Mismatches")
    ws_desc.append(issue_headers)
    for cell in ws_desc[1]:
        cell.font = Font(bold=True)
    for row in all_rows:
        if "Description" in row["issue"]:
            ws_desc.append(
                [
                    row["severity"],
                    row["issue"],
                    row["endpoint"],
                    row["operationId"],
                    row["mqtt_command"],
                    row["side"],
                    row["field"],
                    row.get("rest", ""),
                    row.get("mqtt", ""),
                    row.get("detail", ""),
                ]
            )
    autosize_columns(ws_desc)

    # Notes / missing schemas
    ws_notes = wb.create_sheet("Missing or Skipped")
    ws_notes.append(["Endpoint", "operationId", "MQTT command", "Note"])
    for cell in ws_notes[1]:
        cell.font = Font(bold=True)
    for note in notes:
        ws_notes.append([note.get("endpoint", ""), note.get("operationId", ""), note.get("mqtt_command", ""), note.get("note", "")])
    autosize_columns(ws_notes)

    wb.save(OUTPUT)


def main() -> None:
    spec = load_yaml(REST_SPEC)
    mapping = json.loads(MAP_JSON.read_text(encoding="utf-8"))
    ops = build_rest_operations(spec, mapping)

    all_rows: list[dict[str, Any]] = []
    notes: list[dict[str, Any]] = []
    alignment_rows: list[dict[str, Any]] = []

    for op in ops:
        endpoint = f"{op['method']} {op['path']}"
        rows, note, align = compare_operation(op, spec)
        all_rows.extend(rows)
        alignment_rows.extend(align)
        if note:
            notes.append(
                {
                    "endpoint": endpoint,
                    "operationId": op["operationId"],
                    "mqtt_command": op["mqtt_command"],
                    "note": note,
                }
            )

    write_report(all_rows, ops, notes, alignment_rows)
    aligned = sum(1 for r in alignment_rows if r["status"] in ("OK", "OK (MQTT path-param field)"))
    print(f"Compared {len(ops)} REST operations")
    print(f"Fields compared: {len(alignment_rows)} ({aligned} aligned)")
    print(f"Field issues: {len(all_rows)}")
    print(f"Notes/skipped: {len(notes)}")
    print(f"Report written to {OUTPUT}")


if __name__ == "__main__":
    main()
