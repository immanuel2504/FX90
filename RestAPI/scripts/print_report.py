#!/usr/bin/env python3
"""Print REST vs MQTT comparison report to stdout."""
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
REPORT = ROOT / "RestAPI" / "rest_vs_mqtt_field_report.xlsx"

wb = openpyxl.load_workbook(REPORT, read_only=True)

print("=" * 70)
print("REST vs MQTT FIELD ALIGNMENT REPORT")
print("=" * 70)
print()
print("--- SUMMARY ---")
for r in wb["Summary"].iter_rows(min_row=1, max_row=25, values_only=True):
    if r[0]:
        print(f"{str(r[0]):<42} {r[1]}")

print()
print("--- ENDPOINT SUMMARY ---")
print(f"{'Endpoint':<45} {'Fields':>6} {'OK':>4} {'Path':>4} {'MM':>3} {'Status':>10}")
print("-" * 75)
for r in wb["Endpoint Summary"].iter_rows(min_row=2, values_only=True):
    print(f"{r[0]:<45} {r[3]:>6} {r[4]:>4} {r[5]:>4} {r[6]:>3} {r[11]:>10}")

print()
print("--- ALL ISSUES ---")
issues = [r for r in wb["All Issues"].iter_rows(min_row=2, values_only=True) if r[0]]
if not issues:
    print("(none — 0 field issues)")
else:
    for r in issues:
        print(r)

print()
print("--- ENUM MISMATCHES ---")
enum_rows = [r for r in wb["Enum Mismatches"].iter_rows(min_row=2, values_only=True) if r[0]]
print("(none)" if not enum_rows else "")
for r in enum_rows:
    print(r)

print()
print("--- DESCRIPTION MISMATCHES ---")
desc_rows = [r for r in wb["Description Mismatches"].iter_rows(min_row=2, values_only=True) if r[0]]
print("(none)" if not desc_rows else "")
for r in desc_rows:
    print(r)

print()
print("--- MISSING / SKIPPED ---")
for r in wb["Missing or Skipped"].iter_rows(min_row=2, values_only=True):
    if r[0]:
        print(r[0])
        if r[3]:
            print(f"  {r[3]}")

print()
print("--- EXPECTED MQTT PATH-PARAM FIELDS ---")
for r in wb["Field Alignment"].iter_rows(min_row=2, values_only=True):
    if r[18] == "OK (MQTT path-param field)":
        desc = (r[13] or "")[:100]
        print(f"  {r[0]} | field={r[3]} | mqtt_required={r[16]}")
        print(f"    {desc}")

print()
print(f"Full Excel report: {REPORT}")
